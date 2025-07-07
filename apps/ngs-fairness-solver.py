import marimo

__generated_with = "0.14.10"
app = marimo.App()


@app.cell
def _():
    import pulp
    import pandas as pd
    import os,io
    from pathlib import Path
    return Path, io, os, pd


@app.cell
def _(mo):
    title = 'NGS Fairness Solver'

    md = mo.md(text="""
    ## Model Overview: Fair Sparse Deployment with Proportional Fairness

    **Objective:**  
    Optimize staff deployment across periods, watches, and divisions to:

    - Meet period-by-period demand for each rank (SA, A)
    - Achieve proportional fairness by division and watch (configurable targets in input excel)
    - Minimize concentration (sparsity) of assignments
    - Balance fairness priorities via weighted penalties

    **Decision Variables:**

    - `x[rank, watch, division, period]`: Number of staff assigned
    - `z[rank, period]`: Max assignments per rank and period (for sparsity)
    - Slack variables: Allow controlled deviations from fairness constraints

    **Key Constraints:**

    - **R1: Demand Satisfaction:**  
      For each period and rank, total assignments = required demand
    - **R2: Division Fairness (all periods):**  
      Each division receives its target share (within tolerance) across all periods
    - **R3: Watch Fairness (all periods):**  
      Each watch receives its target share (within tolerance) across all periods
    - **R4: Division-Watch Proportion (all periods):**  
      Each division-watch pair aligns with combined targets (soft constraint, penalized by slack)
    - **R5: Division Fairness (within each period):**  
      Each division receives its target share of demand in each period (soft constraint, lower penalty)

    **Objective Function:**

    Minimize:

    - Total sparsity (`z` variables)
    - Weighted sum of slack variables for division-watch and per-period division fairness (with configurable weights to prioritize constraints)

    """)

    mo.accordion(items={title: md})
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Load Input File""")
    return


@app.cell
def _():
    # Files base dir
    FILE_DIR = 'apps/files/ngs-fairness-solver-files'
    # Files base dir
    OUTPUT_FILE_DIR = 'apps/files/ngs-fairness-solver-files-output'
    return FILE_DIR, OUTPUT_FILE_DIR


@app.cell
def _(FILE_DIR, Path, mo):
    file_browser = mo.ui.file_browser(
        initial_path=Path(FILE_DIR),
        filetypes=[".xlsx"],
        restrict_navigation=True,
        multiple=False,
        label="Select an existing input file:",
    )


    file_uploader = mo.ui.file(
        filetypes=[".xlsx"], label="Or upload your own input file (.xlsx):"
    )

    file_input_tabs = mo.ui.tabs(
        {
            "Choose Existing": file_browser,
            "Upload New": file_uploader,
        }
    )
    return file_browser, file_input_tabs, file_uploader


@app.cell
def _(file_input_tabs):
    file_input_tabs
    return


@app.cell
def _(FILE_DIR, Path, file_browser, file_uploader, io, mo, pd):
    _excel_file = None
    periods_df = None
    targets_df = None

    if file_uploader.value:
        # Save uploaded file to target directory
        _target_dir = Path(FILE_DIR)
        _target_dir.mkdir(parents=True, exist_ok=True)
        _target_path = _target_dir / file_uploader.name()
        with open(_target_path, "wb") as f:
            f.write(file_uploader.contents())

        _excel_file = pd.ExcelFile(io.BytesIO(file_uploader.contents()))
        _file_status = mo.md(f"🟢 Using uploaded file: **{file_uploader.name()}**")

    elif file_browser.value:
        _excel_file = pd.ExcelFile(file_browser.path())
        _file_status = mo.md(
            f"🟡 Using selected file: **{file_browser.path().name}**"
        )

    else:
        _file_status = mo.md(
            "⚠️ **Please select an existing file or upload your own roster file (.xlsx) to proceed!**"
        )

    if _excel_file is not None:
        # Load period demands
        periods_df = pd.read_excel(_excel_file, sheet_name="Periods")
        # columns: ['period', 'SA', 'A']

        # Load target proportions
        targets_df = pd.read_excel(_excel_file, sheet_name="Target Proportions")
        # columns: ['entity_type', 'entity_code', 'target_proportions']

        # Extract sets
        periods = periods_df["period"].tolist()
        ranks = ["SA", "A"]
        watches = targets_df[targets_df["entity_type"] == "watch"][
            "entity_code"
        ].tolist()
        divisions = targets_df[targets_df["entity_type"] == "division"][
            "entity_code"
        ].tolist()

        # Target proportions
        watch_targets = (
            targets_df[targets_df["entity_type"] == "watch"]
            .set_index("entity_code")["target_proportion"]
            .to_dict()
        )
        division_targets = (
            targets_df[targets_df["entity_type"] == "division"]
            .set_index("entity_code")["target_proportion"]
            .to_dict()
        )

        _tabs = mo.ui.tabs(
            {
                "INPUT Sheet": mo.ui.table(periods_df),
                "SHIFTS Sheet": mo.ui.table(targets_df),
            }
        )
        # Display status and data
        data_display = mo.vstack([_file_status, _tabs])
    else:
        data_display = _file_status


    data_display
    return (
        division_targets,
        divisions,
        periods,
        periods_df,
        ranks,
        targets_df,
        watch_targets,
        watches,
    )


@app.cell
def _(mo):
    # Tolerance for overall ratio constraints (in %)
    TOLERANCE_DIV = mo.ui.number(
        start=0,
        stop=100,
        step=1,
        value=1,
        label="Tolerance for overall division ratio constraints (in %)",
    )


    TOLERANCE_WATCH = mo.ui.number(
        start=0,
        stop=100,
        step=1,
        value=0,
        label="Tolerance for overall watch ratio constraints (in %)",
    )

    mo.vstack(items=[TOLERANCE_DIV, TOLERANCE_WATCH])

    return TOLERANCE_DIV, TOLERANCE_WATCH


@app.cell
def _(
    TOLERANCE_DIV,
    TOLERANCE_WATCH,
    division_targets,
    divisions,
    mo,
    periods,
    periods_df,
    ranks,
    targets_df,
    watch_targets,
    watches,
):
    from ortools.sat.python import cp_model


    model = cp_model.CpModel()

    if periods_df is not None and targets_df is not None:
        # Decision variables: x[rank, w, d, p]
        x = {}
        for rank in ranks:
            for w in watches:
                for d in divisions:
                    for p in periods:
                        x[rank, w, d, p] = model.NewIntVar(0, cp_model.INT32_MAX, f'x_{rank}_{w}_{d}_{p}')

        # Auxiliary variables for sparsity: z[rank, p]
        z = {}
        for rank in ranks:
            for p in periods:
                z[rank, p] = model.NewIntVar(0, cp_model.INT32_MAX, f'z_{rank}_{p}')

        # Slack variables for soft constraints
        slack_divwatch_pos, slack_divwatch_neg = {}, {}
        for rank in ranks:
            for d in divisions:
                for w in watches:
                    slack_divwatch_pos[rank, d, w] = model.NewIntVar(0, cp_model.INT32_MAX, f'slack_divwatch_pos_{rank}_{d}_{w}')
                    slack_divwatch_neg[rank, d, w] = model.NewIntVar(0, cp_model.INT32_MAX, f'slack_divwatch_neg_{rank}_{d}_{w}')

        # Hard constraint: meet period demand for each rank
        for idx, p in enumerate(periods):
            for rank in ranks:
                period_demand = int(periods_df.loc[idx, rank])
                model.Add(
                    sum(x[rank, w, d, p] for w in watches for d in divisions) == period_demand
                )

        # Hard constraints: assignments to each watch within tolerance of targets for each rank
        for rank in ranks:
            total_rank = int(periods_df[rank].sum())
            for w in watches:
                target = int(round(watch_targets[w] * total_rank))
                tol = int(round(TOLERANCE_DIV.value * 0.01 * total_rank))
                actual = sum(x[rank, w, d, p] for d in divisions for p in periods)
            
                model.Add(actual >= target - tol)
                model.Add(actual <= target + tol)

        # Hard constraints: assignments to each division within tolerance of targets for each rank
        for rank in ranks:
            total_rank = int(periods_df[rank].sum())
            for d in divisions:
                target = int(round(division_targets[d] * total_rank))
                tol = int(round(TOLERANCE_WATCH.value * 0.01 * total_rank))
                actual = sum(x[rank, w, d, p] for w in watches for p in periods)
                model.Add(actual >= target - tol)
                model.Add(actual <= target + tol)

        # Sparsity constraint: no (watch, division) assignment exceeds z[rank, p]
        for rank in ranks:
            for p in periods:
                for w in watches:
                    for d in divisions:
                        model.Add(x[rank, w, d, p] <= z[rank, p])

        # Soft constraints: intra-division watch ratio deviations
        for rank in ranks:
            total_rank = int(periods_df[rank].sum())
            for d in divisions:
                total_div = int(round(division_targets[d] * total_rank))
                for w in watches:
                    target = int(round(watch_targets[w] * total_div))
                    actual = sum(x[rank, w, d, p] for p in periods)
                    model.Add(actual - target <= slack_divwatch_pos[rank, d, w])
                    model.Add(target - actual <= slack_divwatch_neg[rank, d, w])

        # Slack variables for intra-division period ratios
        slack_divperiod_pos, slack_divperiod_neg = {}, {}
        for rank in ranks:
            for d in divisions:
                for p in periods:
                    slack_divperiod_pos[rank, d, p] = model.NewIntVar(0, cp_model.INT32_MAX, f'slack_divperiod_pos_{rank}_{d}_{p}')
                    slack_divperiod_neg[rank, d, p] = model.NewIntVar(0, cp_model.INT32_MAX, f'slack_divperiod_neg_{rank}_{d}_{p}')

        # Soft constraints: intra-division period ratio deviations
        for idx, p in enumerate(periods):
            for rank in ranks:
                period_demand = int(periods_df.loc[idx, rank])
                for d in divisions:
                    target = int(round(division_targets[d] * period_demand))
                    actual = sum(x[rank, w, d, p] for w in watches)
                    model.Add(actual - target <= slack_divperiod_pos[rank, d, p])
                    model.Add(target - actual <= slack_divperiod_neg[rank, d, p])

        # Slack variables for inter-division watch ratio per period
        slack_interdivwatch_pos, slack_interdivwatch_neg = {}, {}
        for rank in ranks:
            for w in watches:
                for p in periods:
                    slack_interdivwatch_pos[rank, w, p] = model.NewIntVar(0, cp_model.INT32_MAX, f'slack_interdivwatch_pos_{rank}_{w}_{p}')
                    slack_interdivwatch_neg[rank, w, p] = model.NewIntVar(0, cp_model.INT32_MAX, f'slack_interdivwatch_neg_{rank}_{w}_{p}')

        # Soft constraints: inter-division watch ratio per period
        for rank in ranks:
            for w in watches:
                for idx, p in enumerate(periods):
                    period_demand = int(periods_df.loc[idx, rank])
                    target = int(round(watch_targets[w] * period_demand))
                    actual = sum(x[rank, w, d, p] for d in divisions)
                    model.Add(actual - target <= slack_interdivwatch_pos[rank, w, p])
                    model.Add(target - actual <= slack_interdivwatch_neg[rank, w, p])

        # Objective function: minimize sparsity and penalize soft constraint violations (scale slack terms back down)
        OBJ_SCALE = 1000  # or any large integer for precision

        weight_sparsity = 1 * OBJ_SCALE         # 1000
        weight_divwatch = int(0.1 * OBJ_SCALE)  # 100
        weight_divperiod = int(0.5 * OBJ_SCALE)  # 50
        weight_interdivwatch = int(0.3 * OBJ_SCALE)  # 10
    
        model.Minimize(
            weight_sparsity * sum(z[rank, p] for rank in ranks for p in periods)
            + weight_divwatch * sum(slack_divwatch_pos[rank, d, w] + slack_divwatch_neg[rank, d, w]
                                    for rank in ranks for d in divisions for w in watches)
            + weight_divperiod * sum(slack_divperiod_pos[rank, d, p] + slack_divperiod_neg[rank, d, p]
                                     for rank in ranks for d in divisions for p in periods)
            + weight_interdivwatch * sum(slack_interdivwatch_pos[rank, w, p] + slack_interdivwatch_neg[rank, w, p]
                                         for rank in ranks for w in watches for p in periods)
        )

        # Solve the model
        solver = cp_model.CpSolver()
        status = solver.Solve(model)

        mo.output.replace(f'Solver status: {solver.StatusName(status)}')

    return cp_model, solver, status, x


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Display Results""")
    return


@app.cell
def _():
    ordered_divisions = ['HOS', 'KWL', 'KCE', 'NTE', 'NTW']
    ordered_watches = ['A', 'B', 'C', 'D', 'E']
    return ordered_divisions, ordered_watches


@app.cell
def _(
    cp_model,
    division_targets,
    divisions,
    mo,
    ordered_divisions,
    ordered_watches,
    pd,
    periods,
    ranks,
    solver,
    status,
    watch_targets,
    watches,
    x,
):
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        deployment_data = []
        for rank_1 in ranks:
            for p_1 in periods:
                for w_1 in watches:
                    for d_1 in divisions:
                        val = solver.Value(x[rank_1, w_1, d_1, p_1])
                        deployment_data.append({'period': p_1, 'rank': rank_1, 'watch': w_1, 'division': d_1, 'count': val})
        results_df = pd.DataFrame(deployment_data)

        tab_dict = {}
        for rank_1 in ranks:
            rank_df = results_df[results_df['rank'] == rank_1]
            total = rank_df['count'].sum()
    
            # --- Quick Summary ---
            summary_table = mo.ui.table(
                pd.DataFrame({
                    "Total Deployments": [total],
                    "Periods": [len(periods)],
                    "Divisions": [len(divisions)],
                    "Watches": [len(watches)],
                }).T.rename(columns={0: "Value"})
            )
    
            # --- Deployments by Period (as nested tabs) ---
            period_tabs = {}
            for p_1 in periods:
                period_df = rank_df[rank_df['period'] == p_1]
                grid = pd.DataFrame(0, index=ordered_divisions, columns=ordered_watches)
                for _, row in period_df.iterrows():
                    grid.at[row['division'], row['watch']] = row['count']
                grid.loc['TOTAL'] = grid.sum()
                grid['TOTAL'] = grid.sum(axis=1)
                period_tabs[f"Period {p_1}"] = mo.vstack([
                    mo.md(f"#### Deployments for Period {p_1}"),
                    mo.ui.table(grid)
                ])
            deployments_by_period = mo.ui.tabs(period_tabs)
    
            # --- Overall Proportions ---
            watch_actual = rank_df.groupby('watch')['count'].sum() / total
            div_actual = rank_df.groupby('division')['count'].sum() / total
    
            # Watch proportions
            watch_prop_df = pd.DataFrame({
                "Actual": watch_actual,
                "Target": pd.Series(watch_targets),
            })
            watch_prop_df["Delta"] = watch_prop_df["Actual"] - watch_prop_df["Target"]
    
            # Format as percentage strings for display (do NOT use pd.to_numeric on percent strings)
            watch_prop_df_fmt = watch_prop_df.copy()
            for col in ["Actual", "Target", "Delta"]:
                watch_prop_df_fmt[col] = watch_prop_df_fmt[col].apply(lambda x: f"{x:+.2%}" if col == "Delta" else f"{x:.2%}")
    
            # Division proportions
            div_prop_df = pd.DataFrame({
                "Actual": div_actual,
                "Target": pd.Series(division_targets),
            })
            div_prop_df["Delta"] = div_prop_df["Actual"] - div_prop_df["Target"]
    
            div_prop_df_fmt = div_prop_df.copy()
            for col in ["Actual", "Target", "Delta"]:
                div_prop_df_fmt[col] = div_prop_df_fmt[col].apply(lambda x: f"{x:+.2%}" if col == "Delta" else f"{x:.2%}")
    
            # --- Intra-Division Watch Ratio Summary (as nested tabs) ---
            intra_tabs = {}
            for d_1 in divisions:
                div_df = rank_df[rank_df['division'] == d_1]
                total_in_div = div_df['count'].sum()
                summary = []
                for w_1 in watches:
                    actual_1 = div_df[div_df['watch'] == w_1]['count'].sum()
                    actual_prop = actual_1 / total_in_div if total_in_div > 0 else 0
                    target_prop = watch_targets[w_1]
                    delta = actual_prop - target_prop
                    summary.append({
                        'Watch': w_1,
                        'Actual': f'{actual_prop:.2%}',
                        'Target': f'{target_prop:.2%}',
                        'Delta': f'{delta:+.2%}'
                    })
                summary_df = pd.DataFrame(summary).set_index('Watch')
                intra_tabs[f"Division {d_1}"] = mo.vstack([
                    mo.md(f"**Total assigned:** {total_in_div}"),
                    mo.ui.table(summary_df)
                ])
            intra_division_ratios = mo.ui.tabs(intra_tabs)
    
            # --- Compose all content for this rank ---
            tab_dict[rank_1] = mo.vstack([
                mo.md(f"# {rank_1} Deployment Summary"),
                summary_table,
                mo.md("## Deployments by Period"),
                deployments_by_period,
                mo.md("## Overall Proportions"),
                mo.md("**Watch Proportions**"),
                mo.ui.table(watch_prop_df_fmt),
                mo.md("**Division Proportions**"),
                mo.ui.table(div_prop_df_fmt),
                mo.md("## Intra-Division Watch Ratio Summary"),
                intra_division_ratios
            ])
    
        mo.output.clear()
        mo.output.append(mo.ui.tabs(tab_dict))
    else:
        mo.output.clear()
        mo.output.append(mo.md('No optimal solution found. Try increasing TOLERANCE.'))

    return (results_df,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Save Result to xlsx""")
    return


@app.cell
def _(
    OUTPUT_FILE_DIR,
    cp_model,
    mo,
    ordered_divisions,
    ordered_watches,
    os,
    pd,
    periods,
    ranks,
    results_df,
    status,
):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        output_file_name = "deployment_report.xlsx"
        output_excel = os.path.join(OUTPUT_FILE_DIR, output_file_name)
        with pd.ExcelWriter(output_excel, engine="openpyxl") as _writer:
            # Write deployment matrices and proportion tables
            for _p in periods:
                for _rank in ranks:
                    _rank_df = results_df[results_df["rank"] == _rank]
                    _period_df = _rank_df[_rank_df["period"] == _p]
                    _grid = pd.DataFrame(
                        0, index=ordered_divisions, columns=ordered_watches
                    )
                    for _, _row in _period_df.iterrows():
                        _grid.at[_row["division"], _row["watch"]] = _row["count"]
                    _grid.loc["TOTAL"] = _grid.sum()
                    _grid["TOTAL"] = _grid.sum(axis=1)
                    _sheet_name = f"{_rank}_{_p}"
                    _grid.to_excel(_writer, sheet_name=_sheet_name)
                _total = _rank_df["count"].sum()

            # --- Add overall summary sheet ---
            _summary_rows = []
            for _rank in ranks:
                _rank_df = results_df[results_df["rank"] == _rank]
                _total = _rank_df["count"].sum()
                _watch_actual = _rank_df.groupby("watch")["count"].sum() / _total
                _div_actual = _rank_df.groupby("division")["count"].sum() / _total
                _summary_rows.append({"Rank": _rank, "Metric": "Total Deployments", "Value": _total})
                for _watch, _val in _watch_actual.items():
                    _summary_rows.append({"Rank": _rank, "Metric": f"Watch Actual - {_watch}", "Value": f"{_val:.2%}"})
                for _div, _val in _div_actual.items():
                    _summary_rows.append({"Rank": _rank, "Metric": f"Division Actual - {_div}", "Value": f"{_val:.2%}"})

            _summary_df = pd.DataFrame(_summary_rows)
            _summary_df.to_excel(_writer, sheet_name="Overall_Summary", index=False)

        # --- Apply formatting to the Excel file ---
        _wb = load_workbook(output_excel)
        _header_font = Font(bold=True, color="FFFFFF")
        _header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        _center_alignment = Alignment(horizontal="center", vertical="center")

        for _sheet_name in _wb.sheetnames:
            _ws = _wb[_sheet_name]
            # Format header row
            for _cell in _ws[1]:
                _cell.font = _header_font
                _cell.fill = _header_fill
                _cell.alignment = _center_alignment
            # Format index column (if present)
            for _row in _ws.iter_rows(min_row=2, max_row=_ws.max_row, min_col=1, max_col=1):
                for _cell in _row:
                    _cell.font = _header_font
                    _cell.fill = _header_fill
                    _cell.alignment = _center_alignment
            # Auto-fit column widths
            for _col in _ws.columns:
                _max_length = 0
                _col_letter = _col[0].column_letter
                for _cell in _col:
                    try:
                        if _cell.value:
                            _max_length = max(_max_length, len(str(_cell.value)))
                    except Exception:
                        pass
                _ws.column_dimensions[_col_letter].width = _max_length + 2

        _wb.save(output_excel)

        mo.output.append(
            mo.md(f"All deployment matrices, proportion tables, and overall summary exported to **{output_excel}** with enhanced formatting."
        ))
    else:
        mo.output.append("No optimal solution found. Try increasing TOLERANCE.")
    return output_excel, output_file_name


@app.cell
def _(mo, output_excel, output_file_name):
    mo.download(
        data=lambda: open(output_excel, "rb"),
        filename=output_file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        label="Download Output file"
    )
    return


@app.cell
def _():
    import marimo as mo
    return (mo,)


if __name__ == "__main__":
    app.run()
